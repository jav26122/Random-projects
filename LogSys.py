

import os, subprocess, shutil, time, datetime
from os import path
from openpyxl import Workbook
from openpyxl import load_workbook

total = 0
runningAvg = 0
count = 1
loss = 0

if not path.exists("Log.xlsx"):
    book = Workbook()
    sheet = book.active
    sheet["A1"] = "Time"
    sheet["B1"] = "ping (ms)"
    sheet["C1"] = "average (ms)"
    book.save(filename="Log.xlsx")

book = load_workbook(filename="Log.xlsx")
sheet = book.active




while True:
    
    data = []
    time = str(datetime.datetime.now())
    data.append(time)
    try:
        ping = subprocess.Popen("ping google.com -n 1", stdout=subprocess.PIPE)
        output = ping.communicate()
        output = str(output)
        timems = output.find("time=")
        timemsend = output.find("ms", timems, timems + 14)
        timems = output[timems+5:timemsend+2]
        timemsint = timems[0:len(timems) - 2]
        timemsint = float(timemsint)
        total = total + timemsint
        runningAvg = total / count
        print("Average: "+str(runningAvg)+"ms")
        data.append(timemsint)
        data.append(runningAvg)
        sheet.append(data)
        count = count + 1
    except:
        print("No response from host")
        loss = loss + 1


    
    
    

    book.save(filename = "Log.xlsx")




